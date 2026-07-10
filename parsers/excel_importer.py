from pathlib import Path

from openpyxl import load_workbook

from database.db import connect


def import_excel(file_path: Path, document_type: str):

    workbook = load_workbook(
        filename=file_path,
        data_only=True,
    )

    conn = connect()
    cursor = conn.cursor()

    total_rows = 0

    for sheet in workbook.worksheets:

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):

            text = " | ".join(
                [
                    str(value).strip()
                    for value in row
                    if value is not None
                ]
            )

            if not text:
                continue

            cursor.execute(
                """
                INSERT INTO documents
                (
                    filename,
                    document_type,
                    page,
                    content
                )
                VALUES
                (
                    ?, ?, ?, ?
                )
                """,
                (
                    file_path.name,
                    document_type,
                    row_index,
                    text,
                ),
            )

            total_rows += 1

    conn.commit()
    conn.close()

    return total_rows